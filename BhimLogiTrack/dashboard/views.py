import os
import glob
import xlrd
from openpyxl import load_workbook, Workbook
from django.shortcuts import render, redirect
from dashboard.models import UploadedFile
from django.contrib import messages
from django.conf import settings
from django.core.files.storage import default_storage


def home(request):
    return render(request, 'dashboard/home.html')

def upload_file(request):
    if request.method == 'POST' and request.FILES.getlist('excel_files'):
        input_folder = os.path.join(settings.MEDIA_ROOT, 'inputs')
        os.makedirs(input_folder, exist_ok=True)
    
        for f in os.listdir(input_folder):
            file_path = os.path.join(input_folder, f)
            if os.path.isfile(file_path):
                os.remove(file_path)

        uploaded_files = request.FILES.getlist('excel_files')
        saved_files = []
        tag = request.POST.get('tag', 'other')  # Optional tag for files
        
        for file in uploaded_files:
            try:
                # Save the file to the inputs folder
                file_path = os.path.join(input_folder, file.name)
                with default_storage.open(file_path, 'wb+') as destination:
                    for chunk in file.chunks():
                        destination.write(chunk)
                        UploadedFile.objects.create(file=file, tag=tag)
                        saved_files.append(file.name)
                        messages.success(request, f"File '{file.name}' uploaded successfully!")  # Success message
            except Exception as e:
                messages.error(request, f"Error uploading file '{file.name}': {str(e)}")  # Error message

        if saved_files:
            # Call consolidate method after uploading files
            total_amount, invoice_string = consolidate_invoices_from_loadsheets()
            return render(request, 'dashboard/home.html', {
                'total_amount': total_amount,
                'invoice_string': invoice_string
            })  # Render with the new data
        else:
            messages.error(request, "No valid files uploaded.")  # Error if no valid files are uploaded

    messages.error(request, "No files uploaded.")  # Error if no files selected
    return render(request, 'dashboard/home.html')  # Adjust this path if needed
    
def consolidate_invoices_from_loadsheets():
    input_folder = 'media/inputs'
    output_file = 'media/outputs/invoices_with_amounts.xlsx'
    invoices = set()

    # Extract from .xlsx files
    for file_path in glob.glob(os.path.join(input_folder, '*.xlsx')):
        print(f'Processing XLSX: {file_path}')
        wb = load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for i, cell in enumerate(row):
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("T00"):
                        amount = row[i + 1].value if i + 1 < len(row) else 0
                        invoices.add((cell.value, float(amount) if amount else 0))

    # Extract from .xls files
    for file_path in glob.glob(os.path.join(input_folder, '*.xls')):
        print(f'Processing XLS: {file_path}')
        wb = xlrd.open_workbook(file_path)
        for sheet in wb.sheets():
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                for i, cell_value in enumerate(row):
                    if isinstance(cell_value, str) and cell_value.startswith("T00"):
                        amount = row[i + 1] if i + 1 < len(row) else 0
                        invoices.add((cell_value, float(amount) if amount else 0))

    # Save to Excel file
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = 'Invoices'

    # Write headers
    output_ws.cell(row=1, column=1, value='Invoice Number')
    output_ws.cell(row=1, column=2, value='Amount')

    # Write invoice data and calculate totals
    total_amount = 0.0
    invoice_list = []

    for idx, (invoice, amount) in enumerate(sorted(invoices), start=2):
        output_ws.cell(row=idx, column=1, value=invoice)
        output_ws.cell(row=idx, column=2, value=amount)
        total_amount += amount
        invoice_list.append(f"'{invoice}'")

    # Create invoice string
    invoice_string = ','.join(invoice_list)

    # Save the workbook
    output_wb.save(output_file)

    # Print the results
    print(f"\n✅ Done! Extracted {len(invoices)} invoice records.")
    print(f"📁 Output saved to: {output_file}")
    print(f"🧮 Total Amount: {total_amount:,.2f}")
    print(f"🧾 Invoice String:\n{invoice_string}")
    return total_amount, invoice_string